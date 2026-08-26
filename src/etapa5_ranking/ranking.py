"""
Módulo de Cálculo de Ranking e Scoring MCDA (Etapa 5).
Responsável: Membro 3

Aplica o algoritmo Multi-Critério (MCDA Ponderado) com normalização Min-Max
para ranquear os fornecedores ativos com base em Custo, Prazo, Capacidade e Qualidade.
"""

import os
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
import openpyxl
import pandas as pd

from src.logger import AuditLogger

logger = logging.getLogger("Hyperautomation")

# Pesos padrão conforme PDD e folha da Avaliação 03
DEFAULT_WEIGHTS = {
    "custo": 0.40,
    "prazo": 0.25,
    "capacidade": 0.20,
    "qualidade": 0.15
}


def _get_val(d: dict, *keys, default=None):
    """Obtém valor de um dicionário buscando por múltiplas chaves (case-insensitive)."""
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
    for k in keys:
        for actual_k, v in d.items():
            if actual_k.lower() == k.lower() and v is not None:
                return v
    return default


def normalize_status(status_str: Optional[str]) -> str:
    """Normaliza e valida a string de status para 'ATIVO' ou 'BLOQUEADO'."""
    if not status_str:
        return "ATIVO"  # Padrão caso venha vazio

    status_clean = str(status_str).strip().upper()
    if status_clean in ["BLOQUEADO", "INATIVO", "BLOCKED", "REJEITADO"]:
        return "BLOQUEADO"
    return "ATIVO"


def normalize_inverse_value(value: float, min_val: float, max_val: float) -> float:
    """
    Normalização Min-Max onde menor é melhor (Custo, Prazo).
    Formula: (max - val) / (max - min)
    """
    if max_val == min_val:
        return 1.0
    return (max_val - value) / (max_val - min_val)


def normalize_direct_value(value: float, min_val: float, max_val: float) -> float:
    """
    Normalização Min-Max onde maior é melhor (Capacidade, Qualidade).
    Formula: (val - min) / (max - min)
    """
    if max_val == min_val:
        return 1.0
    return (value - min_val) / (max_val - min_val)


def get_extrema(active_suppliers: List[Dict[str, Any]]) -> Optional[Dict[str, Tuple[float, float]]]:
    """Retorna os valores mínimo e máximo para cada critério considerando fornecedores ativos."""
    if not active_suppliers:
        return None

    custos = [float(_get_val(s, "Custo", "custo", default=0.0)) for s in active_suppliers]
    prazos = [float(_get_val(s, "Prazo_Dias", "prazo_dias", "prazo", default=0.0)) for s in active_suppliers]
    capacidades = [float(_get_val(s, "Capacidade", "capacidade", default=0.0)) for s in active_suppliers]
    qualidades = [float(_get_val(s, "Qualidade", "qualidade", default=0.0)) for s in active_suppliers]

    return {
        "custo": (min(custos), max(custos)),
        "prazo": (min(prazos), max(prazos)),
        "capacidade": (min(capacidades), max(capacidades)),
        "qualidade": (min(qualidades), max(qualidades)),
    }


def calculate_individual_score(
    supplier: Dict[str, Any],
    weights: Dict[str, float],
    extrema: Dict[str, Tuple[float, float]]
) -> Tuple[float, Dict[str, float]]:
    """
    Calcula a nota ponderada final e os scores parciais de um fornecedor.

    Returns:
        Tupla (nota_final, scores_parciais)
    """
    custo = float(_get_val(supplier, "Custo", "custo", default=0.0))
    prazo = float(_get_val(supplier, "Prazo_Dias", "prazo_dias", "prazo", default=0.0))
    capacidade = float(_get_val(supplier, "Capacidade", "capacidade", default=0.0))
    qualidade = float(_get_val(supplier, "Qualidade", "qualidade", default=0.0))

    s_custo = normalize_inverse_value(custo, *extrema["custo"])
    s_prazo = normalize_inverse_value(prazo, *extrema["prazo"])
    s_capacidade = normalize_direct_value(capacidade, *extrema["capacidade"])
    s_qualidade = normalize_direct_value(qualidade, *extrema["qualidade"])

    score = (
        s_custo * weights["custo"]
        + s_prazo * weights["prazo"]
        + s_capacidade * weights["capacidade"]
        + s_qualidade * weights["qualidade"]
    )

    scores_parciais = {
        "score_custo": round(s_custo, 4),
        "score_prazo": round(s_prazo, 4),
        "score_capacidade": round(s_capacidade, 4),
        "score_qualidade": round(s_qualidade, 4)
    }

    return score, scores_parciais


def extrair_pesos_criterios(df_criterios: Any) -> Dict[str, float]:
    """Extrai os pesos dos critérios a partir de um DataFrame ou dicionário, com fallback padrão."""
    pesos = DEFAULT_WEIGHTS.copy()
    if df_criterios is None:
        return pesos

    try:
        if isinstance(df_criterios, pd.DataFrame):
            for _, row in df_criterios.iterrows():
                crit = str(row.get("Criterio", "")).strip().lower()
                peso = float(row.get("Peso", 0))
                if "custo" in crit:
                    pesos["custo"] = peso
                elif "prazo" in crit:
                    pesos["prazo"] = peso
                elif "capacidade" in crit:
                    pesos["capacidade"] = peso
                elif "qualidade" in crit:
                    pesos["qualidade"] = peso
        elif isinstance(df_criterios, dict):
            for k, v in df_criterios.items():
                crit = str(k).strip().lower()
                if "custo" in crit:
                    pesos["custo"] = float(v)
                elif "prazo" in crit:
                    pesos["prazo"] = float(v)
                elif "capacidade" in crit:
                    pesos["capacidade"] = float(v)
                elif "qualidade" in crit:
                    pesos["qualidade"] = float(v)
    except Exception as e:
        logger.warning(f"[RANKING] Falha ao extrair pesos de critérios: {e}. Utilizando pesos padrão.")

    return pesos


def calculate_ranking(
    suppliers: List[Dict[str, Any]],
    weights: Optional[Dict[str, float]] = None
) -> List[Dict[str, Any]]:
    """Calcula scores para fornecedores ativos e deixa bloqueados com score None."""
    if weights is None:
        weights = DEFAULT_WEIGHTS.copy()

    # Cria cópia rasa dos dicts para não mutar os originais inesperadamente
    suppliers_copy = [s.copy() for s in suppliers]

    for s in suppliers_copy:
        raw_status = _get_val(s, "Status", "status", default="ATIVO")
        s["status_normalizado"] = normalize_status(raw_status)

    active_suppliers = [s for s in suppliers_copy if s["status_normalizado"] == "ATIVO"]
    extrema = get_extrema(active_suppliers)

    for s in suppliers_copy:
        if s["status_normalizado"] == "BLOQUEADO":
            s["score"] = None
        else:
            if extrema:
                score, _ = calculate_individual_score(s, weights, extrema)
                s["score"] = score
            else:
                s["score"] = 0.0

    return sorted(
        suppliers_copy,
        key=lambda x: (x["score"] is not None, x["score"] if x["score"] is not None else -1),
        reverse=True
    )


def calcular_ranking_ponderado(
    propostas: Union[List[Dict[str, Any]], pd.DataFrame],
    df_criterios: Any = None,
    audit: Optional[AuditLogger] = None,
    output_path: Optional[Union[str, Path]] = None
) -> pd.DataFrame:
    """
    Função principal de integração da Etapa 5.
    Calcula o ranking ponderado MCDA, registra a auditoria e retorna um DataFrame padronizado.
    """
    logger.info("[RANKING] Iniciando Etapa 5 - Cálculo do Ranking Ponderado MCDA")

    if isinstance(propostas, pd.DataFrame):
        lista_propostas = propostas.to_dict(orient="records")
    else:
        lista_propostas = list(propostas)

    pesos = extrair_pesos_criterios(df_criterios)
    logger.info(f"[RANKING] Pesos aplicados: {pesos}")

    ranking_ordenado = calculate_ranking(lista_propostas, weights=pesos)

    active_suppliers = [s for s in ranking_ordenado if s.get("status_normalizado") == "ATIVO"]
    extrema = get_extrema(active_suppliers)

    registros_finais = []
    for index, s in enumerate(ranking_ordenado, 1):
        fornecedor = str(_get_val(s, "Fornecedor", "fornecedor", default="N/A"))
        score = s.get("score")
        nota_final = round(score, 4) if score is not None else None
        status = s.get("status_normalizado", "ATIVO")
        observacao = str(_get_val(s, "Observacao", "observacao", default=""))

        # Registro no audit logger
        if audit is not None:
            if score is not None and extrema:
                _, scores_parciais = calculate_individual_score(s, pesos, extrema)
                audit.registrar_calculo(fornecedor, scores_parciais, nota_final)
            elif score is not None:
                audit.registrar_calculo(fornecedor, {}, nota_final)

        registros_finais.append({
            "Posicao": index,
            "Fornecedor": fornecedor,
            "Nota_Final": nota_final,
            "Status": status,
            "Observacao": observacao
        })

    df_ranking = pd.DataFrame(registros_finais)

    if audit is not None:
        audit.registrar_ranking(df_ranking.to_dict(orient="records"))

    if output_path:
        fill_spreadsheet(ranking_ordenado, filename=str(output_path))

    logger.info(f"[RANKING] Etapa 5 concluída com sucesso. Total ranqueados: {len(df_ranking)}")
    return df_ranking


def fill_spreadsheet(suppliers: List[Dict[str, Any]], filename: str = "modelo_ranking.xlsx") -> None:
    """Calcula o ranking e salva na planilha Excel."""
    if os.path.isabs(filename):
        target_path = Path(filename)
    else:
        base_dir = Path(__file__).resolve().parent
        target_path = (base_dir / ".." / ".." / "resources" / "01_SELECAO_FORNECEDORES" / filename).resolve()

    target_path.parent.mkdir(parents=True, exist_ok=True)
    ranking = calculate_ranking(suppliers)

    if target_path.exists():
        wb = openpyxl.load_workbook(target_path)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Posicao", "Fornecedor", "Nota_Final", "Status", "Observacao"])

    for index, s in enumerate(ranking, 1):
        posicao = index
        fornecedor = str(_get_val(s, "Fornecedor", "fornecedor", default="N/A"))
        nota_final = round(s["score"], 4) if s.get("score") is not None else None
        status = s.get("status_normalizado", "ATIVO")
        observacao = str(_get_val(s, "Observacao", "observacao", default=""))

        ws.append([posicao, fornecedor, nota_final, status, observacao])

    wb.save(target_path)
    logger.info(f"[RANKING] Planilha salva com sucesso em: {target_path}")
