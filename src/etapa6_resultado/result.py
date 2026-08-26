"""
Módulo de Geração de Resultados e Relatórios (Etapa 6).
Responsável: Membro 3

Consolida a tabela final de classificação (ativos ranqueados + rejeitados/bloqueados),
preenche a planilha Excel oficial e gera o Dashboard HTML executivo via Jinja2.
"""

import os
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
import openpyxl
import pandas as pd
from jinja2 import Environment, FileSystemLoader

from src.logger import AuditLogger

logger = logging.getLogger("Hyperautomation")


def load_ranking_data(filename: str = "modelo_ranking.xlsx") -> List[Dict[str, Any]]:
    """Lê os dados de ranking gravados na planilha Excel."""
    if os.path.isabs(filename):
        excel_path = Path(filename)
    else:
        base_dir = Path(__file__).resolve().parent
        excel_path = (base_dir / ".." / ".." / "resources" / "01_SELECAO_FORNECEDORES" / filename).resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        pos_val = row[0] if row[0] is not None else "-"
        forn_val = str(row[1]) if row[1] is not None else "Sem Nome"
        status_val = str(row[3]).strip() if row[3] is not None else "Pendente"
        is_blocked = status_val.upper() in ["BLOQUEADO", "INATIVO", "BLOCKED", "REJEITADO"]

        # Se estiver bloqueado ou a nota for None, atribui "-"
        if is_blocked or row[2] is None or row[2] == "-":
            nota_final = "-"
        else:
            try:
                nota_final = round(float(row[2]), 4)
            except (ValueError, TypeError):
                nota_final = row[2]

        obs_val = str(row[4]) if len(row) > 4 and row[4] is not None else "Sem observações"

        data.append({
            "posicao": pos_val,
            "fornecedor": forn_val,
            "nota_final": nota_final,
            "status": status_val,
            "observacao": obs_val
        })

    return data


def generate_html_report(
    data: List[Dict[str, Any]],
    template_name: str = "template_result.html",
    output_file: Union[str, Path] = "relatorio_ranking.html"
) -> str:
    """Gera o relatório HTML carregando o template Jinja2."""
    base_dir = Path(__file__).resolve().parent

    env = Environment(loader=FileSystemLoader(str(base_dir)))
    template = env.get_template(template_name)

    total_suppliers = len(data)

    # Contagens de status
    approved_count = sum(1 for item in data if str(item.get("status", "")).strip().upper() == "ATIVO")
    blocked_count = sum(
        1 for item in data
        if str(item.get("status", "")).strip().upper() in ["BLOQUEADO", "INATIVO", "BLOCKED", "REJEITADO"]
    )

    # Filtra apenas os ativos para encontrar o melhor fornecedor
    active_suppliers = [item for item in data if str(item.get("status", "")).strip().upper() == "ATIVO"]

    best_supplier = None
    if active_suppliers:
        best_supplier = min(
            active_suppliers,
            key=lambda x: x["posicao"] if isinstance(x.get("posicao"), (int, float)) else float("inf")
        )

    html_content = template.render(
        data=data,
        total_suppliers=total_suppliers,
        approved_count=approved_count,
        blocked_count=blocked_count,
        best_supplier=best_supplier
    )

    if os.path.isabs(str(output_file)):
        out_path = Path(output_file)
    else:
        out_path = base_dir / output_file

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    logger.info(f"[RESULTADO] Relatório HTML gerado com sucesso em: {out_path}")
    return str(out_path)


def gerar_resultado_final(
    df_ranking: Union[pd.DataFrame, List[Dict[str, Any]]],
    propostas_rejeitadas: Optional[List[Dict[str, Any]]] = None,
    output_path: Optional[Union[str, Path]] = None,
    audit: Optional[AuditLogger] = None
) -> pd.DataFrame:
    """
    Função principal da Etapa 6.
    Consolida fornecedores ranqueados e rejeitados, gera Excel e relatório HTML.
    """
    logger.info("[RESULTADO] Iniciando Etapa 6 - Geração de Resultados e Relatórios")

    # 1. Normaliza dados do ranking de ativos
    if isinstance(df_ranking, pd.DataFrame):
        registros_ativos = df_ranking.to_dict(orient="records")
    else:
        registros_ativos = [r.copy() for r in df_ranking]

    registros_consolidados = []
    for r in registros_ativos:
        pos = r.get("Posicao") or r.get("posicao") or len(registros_consolidados) + 1
        forn = r.get("Fornecedor") or r.get("fornecedor") or "N/A"
        nota = r.get("Nota_Final") or r.get("nota_final") or r.get("score")
        if nota is not None and nota != "-":
            try:
                nota = round(float(nota), 4)
            except (ValueError, TypeError):
                pass
        else:
            nota = "-"

        status = str(r.get("Status") or r.get("status") or "ATIVO").strip().upper()
        obs = str(r.get("Observacao") or r.get("observacao") or "Proposta válida e classificada.")

        registros_consolidados.append({
            "posicao": pos,
            "fornecedor": forn,
            "nota_final": nota,
            "status": status,
            "observacao": obs
        })

    # 2. Adiciona fornecedores rejeitados/bloqueados da Etapa 3
    if propostas_rejeitadas:
        for r in propostas_rejeitadas:
            forn = r.get("Fornecedor") or r.get("fornecedor") or "Fornecedor Rejeitado"
            obs = r.get("Observacao") or r.get("observacao") or "Proposta desclassificada na validação."
            status = "BLOQUEADO"

            registros_consolidados.append({
                "posicao": "-",
                "fornecedor": forn,
                "nota_final": "-",
                "status": status,
                "observacao": obs
            })

    # 3. Exporta para a planilha Excel de saída
    if output_path is not None:
        target_excel = Path(output_path).resolve()
    else:
        base_dir = Path(__file__).resolve().parent
        target_excel = (base_dir / ".." / ".." / "output" / "ranking_final.xlsx").resolve()

    target_excel.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranking Final"
    ws.append(["Posicao", "Fornecedor", "Nota_Final", "Status", "Observacao"])

    for item in registros_consolidados:
        nota_excel = item["nota_final"] if item["nota_final"] != "-" else None
        ws.append([
            item["posicao"],
            item["fornecedor"],
            nota_excel,
            item["status"],
            item["observacao"]
        ])

    wb.save(target_excel)
    logger.info(f"[RESULTADO] Planilha de ranking exportada para: {target_excel}")

    # 4. Gera o Dashboard HTML
    output_html = target_excel.parent / "relatorio_ranking.html"
    generate_html_report(registros_consolidados, output_file=output_html)

    # 5. Converte para DataFrame final formatado para retorno
    df_resultado = pd.DataFrame([
        {
            "Posicao": item["posicao"],
            "Fornecedor": item["fornecedor"],
            "Nota_Final": item["nota_final"],
            "Status": item["status"],
            "Observacao": item["observacao"]
        }
        for item in registros_consolidados
    ])

    if audit is not None:
        audit.registrar_ranking(df_resultado.to_dict(orient="records"))

    logger.info(f"[RESULTADO] Processamento final concluído com {len(df_resultado)} fornecedores registrados.")
    return df_resultado
